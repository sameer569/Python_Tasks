#importin required Libraries
import os,glob
import xlrd
import openpyxl
import pandas as pd

# Creating an Empty Dictionary
d={}

#Reading text file
f=open("C:\\Users\\Ifra Fatimah\\Desktop\\New folder (3)\\file_1\\file_2\\file_3\\1.txt","r")
line = f.readlines()
data2 = []
for ele in line:
    ele1 = ele.strip()
    data2.append(ele1)

data3 = data2[1:]
for data in data3:
    data4 = data.split(' ')
    d[data4[0]] = data4[1]

print(d)
L1=d.values()
#print(L1)
for l1 in L1:
    print(l1)
L2=d.keys()
for l2 in L2:
    print(l2)


# #print(f.read())
path="C:\\Users\\Ifra Fatimah\\Desktop\\IP_Address.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
r=1
c=1
for l2 in L2:
    sheet.cell(row=r,column=c).value=l2
    r +=1
workbook.save(path)

r=1
c=2
for l1 in L1:
    sheet.cell(row=r,column=c).value=l1
    r +=1
workbook.save(path)